# gesture_temporal.py
# unified temporal gesture pipeline (0..9) with 4 model types:
#   - tcn (default)
#   - lstm
#   - gru
#   - transformer
#
# 0 = no gesture
# 1 = open palm
# 2 = fist
# 3 = thumbs up
# 4 = victory / peace
# 5 = ok sign
# 6 = wave
# 7 = help signal
# 8 = swipe left
# 9 = swipe right

import os, time, uuid, argparse, csv
import numpy as np
import cv2
import mediapipe as mp
from collections import deque
from typing import List

import torch
import torch.nn as nn
import torch.optim as optim
from torch.utils.data import Dataset, DataLoader
from sklearn.model_selection import train_test_split
from sklearn.metrics import confusion_matrix, accuracy_score
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =========================
# Global config
# =========================
mp_hands = mp.solutions.hands
mp_drawing = mp.solutions.drawing_utils
mp_drawing_styles = mp.solutions.drawing_styles

NUM_JOINTS = 21
FEAT_DIM = 128
SEQ_LEN_DEFAULT = 48

NUM_CLASSES = 10
LABEL_MAP = {
    "0": 0,  # no gesture / background
    "1": 1,  # open palm
    "2": 2,  # fist
    "3": 3,  # thumbs up
    "4": 4,  # victory / peace
    "5": 5,  # ok
    "6": 6,  # wave
    "7": 7,  # help
    "8": 8,  # swipe left
    "9": 9,  # swipe right
}

GESTURE_NAMES = [
    "none",         # 0
    "open palm",    # 1
    "fist",         # 2
    "thumbs up",    # 3
    "peace",        # 4
    "OK",           # 5
    "wave",         # 6
    "help signal",  # 7
    "swipe left",   # 8
    "swipe right"   # 9
]

PAD_ZERO = np.zeros((NUM_JOINTS, 3), dtype=np.float32)

import random
def set_global_seed(seed: int = 42):
    import torch.backends.cudnn as cudnn
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)
    cudnn.deterministic = True
    cudnn.benchmark = False
    os.environ["PYTHONHASHSEED"] = str(seed)
    print(f"[seed fixed at {seed}]")

# call it immediately so all random ops are seeded
set_global_seed(42)

# =========================
# Normalization utilities
# =========================
def norm_palm_frame(lm):
    L = np.array([[p.x, p.y, p.z] for p in lm], dtype=np.float32)
    WRIST, INDEX_MCP, MIDDLE_MCP, PINKY_MCP = 0, 5, 9, 17

    O = L[WRIST]
    a = L[INDEX_MCP] - O
    b = L[PINKY_MCP] - O

    u = b - a
    un = np.linalg.norm(u)
    u = u / un if un > 1e-8 else np.array([1, 0, 0], np.float32)

    v = L[MIDDLE_MCP] - O
    v = v - (u @ v) * u
    vn = np.linalg.norm(v)
    v = v / vn if vn > 1e-8 else np.array([0, 1, 0], np.float32)

    w = np.cross(u, v)
    wn = np.linalg.norm(w)
    w = w / wn if wn > 1e-8 else np.array([0, 0, 1], np.float32)

    R = np.stack([u, v, w], axis=1)
    X = (L - O) @ R

    scale = np.linalg.norm(((L[MIDDLE_MCP] - O) @ R)[:2]) + 1e-6
    Xn = X / scale

    mcp_idx = [5, 9, 13, 17]
    if Xn[mcp_idx, 2].mean() < 0:
        Xn[:, 2] *= -1.0

    return Xn


def handpack(frame_hands) -> np.ndarray:
    hands_sorted = {"Left": None, "Right": None}
    for lm, handedness in frame_hands:
        label = handedness.classification[0].label
        hands_sorted[label] = lm

    feats = []
    mask = []
    for side in ["Left", "Right"]:
        if hands_sorted[side] is None:
            feats.append(PAD_ZERO)
            mask.append(0.0)
        else:
            Xn = norm_palm_frame(hands_sorted[side].landmark)
            feats.append(Xn)
            mask.append(1.0)

    F = np.concatenate(feats, axis=0).reshape(-1)
    F = np.concatenate([F, np.array(mask, dtype=np.float32)])
    return F  # (128,)


# =========================
# Models (4 types)
# =========================
class TinyTCN(nn.Module):
    def __init__(self, in_dim=FEAT_DIM, n_classes=NUM_CLASSES):
        super().__init__()
        self.net = nn.Sequential(
            nn.Conv1d(in_dim, 96, 5, padding=2),
            nn.ReLU(inplace=True),
            nn.Conv1d(96, 96, 5, padding=2),
            nn.ReLU(inplace=True),
            nn.AdaptiveAvgPool1d(1)
        )
        self.head = nn.Linear(96, n_classes)

    def forward(self, x):        # [B, T, F]
        x = x.transpose(1, 2)    # -> [B, F, T]
        x = self.net(x).squeeze(-1)
        return self.head(x)


class TinyLSTM(nn.Module):
    def __init__(self, in_dim=FEAT_DIM, hidden=96, n_classes=NUM_CLASSES):
        super().__init__()
        self.lstm = nn.LSTM(in_dim, hidden, batch_first=True)
        self.fc = nn.Linear(hidden, n_classes)

    def forward(self, x):        # [B, T, F]
        out, _ = self.lstm(x)
        out = out[:, -1, :]
        return self.fc(out)


class TinyGRU(nn.Module):
    def __init__(self, in_dim=FEAT_DIM, hidden=96, n_classes=NUM_CLASSES):
        super().__init__()
        self.gru = nn.GRU(in_dim, hidden, batch_first=True)
        self.fc = nn.Linear(hidden, n_classes)

    def forward(self, x):
        out, _ = self.gru(x)
        out = out[:, -1, :]
        return self.fc(out)


class TinyTransformer(nn.Module):
    def __init__(self, in_dim=FEAT_DIM, n_classes=NUM_CLASSES, n_heads=4, depth=2, d_model=128):
        super().__init__()
        self.proj = nn.Linear(in_dim, d_model)
        enc_layer = nn.TransformerEncoderLayer(
            d_model=d_model,
            nhead=n_heads,
            batch_first=True,
            dim_feedforward=256,
            dropout=0.1,
        )
        self.enc = nn.TransformerEncoder(enc_layer, num_layers=depth)
        self.cls = nn.Linear(d_model, n_classes)

    def forward(self, x):        # [B, T, F]
        x = self.proj(x)
        x = self.enc(x)
        x = x.mean(dim=1)
        return self.cls(x)


def make_model(model_type: str, seq_len: int = SEQ_LEN_DEFAULT):
    model_type = model_type.lower()
    if model_type == "tcn":
        return TinyTCN()
    if model_type == "lstm":
        return TinyLSTM()
    if model_type == "gru":
        return TinyGRU()
    if model_type == "transformer":
        return TinyTransformer()
    raise ValueError(f"unknown model_type: {model_type}")


# =========================
# Dataset
# =========================
class SeqDS(Dataset):
    def __init__(self, rows, seq_len=SEQ_LEN_DEFAULT):
        self.rows = rows
        self.seq_len = seq_len

    def __len__(self): return len(self.rows)

    def __getitem__(self, i):
        path, lab = self.rows[i]

        path = str(path).replace("\\", "/")

        # ---- Missing-file handling (NEW) ----
        if not os.path.exists(path):
            print(f"WARNING: missing file, skipping: {path}")
            X = np.zeros((self.seq_len, FEAT_DIM), dtype=np.float32)
            y = 0
            return torch.from_numpy(X), torch.tensor(y, dtype=torch.long)
        # -------------------------------------

        d = np.load(path)
        X = d["seq"].astype(np.float32)
        y = int(d["label"])

        T = X.shape[0]
        if T < self.seq_len:
            pad = np.repeat(X[-1:], self.seq_len - T, axis=0)
            X = np.concatenate([X, pad], axis=0)
        elif T > self.seq_len:
            X = X[-self.seq_len:]

        return torch.from_numpy(X), torch.tensor(y, dtype=torch.long)

def load_index(csv_path):
    rows = []
    with open(csv_path, "r") as f:
        rdr = csv.DictReader(f)
        for r in rdr:
            rows.append((r["path"], int(r["label"])))
    return rows


# =========================
# collect
# =========================
def cmd_collect(args):
    os.makedirs(args.outdir, exist_ok=True)
    index_path = os.path.join(args.outdir, "index.csv")
    if not os.path.exists(index_path):
        with open(index_path, "w") as f:
            f.write("path,label\n")

    print("Collecting... keys: 0=none, 1=palm, 2=fist, 3=thumbs up, 4=peace, 5=ok, 6=wave, 7=help, 8=swipe left, 9=swipe right, q=quit")

    cap = cv2.VideoCapture(0)
    cap.set(cv2.CAP_PROP_FRAME_WIDTH, args.width)
    cap.set(cv2.CAP_PROP_FRAME_HEIGHT, args.height)

    hands = mp_hands.Hands(
        static_image_mode=False,
        max_num_hands=2,
        model_complexity=1,
        min_detection_confidence=0.5,
        min_tracking_confidence=0.5,
    )

    buf: List[np.ndarray] = []

    while True:
        ok, frame = cap.read()
        if not ok:
            break
        if args.flip:
            frame = cv2.flip(frame, 1)

        rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        res = hands.process(rgb)

        pairs = []
        if res.multi_hand_landmarks:
            for lm, hd in zip(res.multi_hand_landmarks, res.multi_handedness):
                pairs.append((lm, hd))
                # draw skeletal landmarks on the BGR frame
                mp_drawing.draw_landmarks(
                    frame,
                    lm,
                    mp_hands.HAND_CONNECTIONS,
                    mp_drawing_styles.get_default_hand_landmarks_style(),
                    mp_drawing_styles.get_default_hand_connections_style(),
                )

        feat = handpack(pairs)
        buf.append(feat)
        if len(buf) > args.seq_len:
            buf.pop(0)

        cv2.putText(frame, f"buffer: {len(buf)}/{args.seq_len}", (12, 28),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 255), 2)
        cv2.putText(frame, "Collecting... keys: 0=none, 1=palm, 2=fist, 3=thumbs up, 4=peace, 5=ok, 6=wave, 7=help, 8=swipe left, 9=swipe right, q=quit", (12, 58),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 1)
        cv2.imshow("collect_temporal", frame)

        k = cv2.waitKey(1) & 0xFF
        if k == ord('q') or k == 27:
            break

        ch = chr(k) if 48 <= k <= 57 else None  # digits 0..9
        if ch in LABEL_MAP and len(buf) == args.seq_len:
            lbl = LABEL_MAP[ch]
            arr = np.stack(buf, axis=0)
            fname = f"{time.strftime('%Y%m%d-%H%M%S')}-{uuid.uuid4().hex[:6]}-L{lbl}.npz"
            path = os.path.join(args.outdir, fname)
            path = str(path).replace("\\", "/")
            np.savez_compressed(path, seq=arr, label=np.int64(lbl))
            with open(index_path, "a") as f:
                f.write(f"{path},{lbl}\n")
            print("saved:", path)

    hands.close()
    cap.release()
    cv2.destroyAllWindows()

def save_metrics_to_excel(metrics_path, model_type, best_epoch, train_metrics_history, val_metrics_history, 
                          test_metrics, confusion_mat):
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  
    
    # Define styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    ws_train = wb.create_sheet("Training Progress", 0)
    ws_train['A1'] = f"Training Metrics - {model_type.upper()}"
    ws_train['A1'].font = Font(bold=True, size=12)
    
    headers = ['Epoch', 'Train Accuracy', 'Val Accuracy', 'Val Macro F1', 'Val Precision', 'Val Recall', 'Best Model']
    for col, header in enumerate(headers, 1):
        cell = ws_train.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
    
    for row, (ep_num, train_dict, val_dict) in enumerate(zip(range(1, len(train_metrics_history)+1), 
                                                               train_metrics_history, 
                                                               val_metrics_history), 4):
        is_best = "✓ BEST" if ep_num == best_epoch else ""
        row_data = [
            ep_num,
            train_dict.get('accuracy', 0),
            val_dict.get('accuracy', 0),
            val_dict.get('macro_f1', 0),
            val_dict.get('macro_precision', 0),
            val_dict.get('macro_recall', 0),
            is_best
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws_train.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = center_align
            if isinstance(value, float):
                cell.number_format = '0.0000'
    
    for col in range(1, len(headers) + 1):
        ws_train.column_dimensions[get_column_letter(col)].width = 15
    
    ws_test_overall = wb.create_sheet("Test Set - Overall", 1)
    ws_test_overall['A1'] = f"Test Set Overall Metrics - {model_type.upper()}"
    ws_test_overall['A1'].font = Font(bold=True, size=12)
    
    test_overall_data = [
        ['Metric', 'Value'],
        ['Overall Accuracy', test_metrics['accuracy']],
        ['Macro F1-Score', test_metrics['macro_f1']],
        ['Macro Precision', test_metrics['macro_precision']],
        ['Macro Recall', test_metrics['macro_recall']],
        ['Best Training Epoch', best_epoch],
    ]
    
    for row, row_data in enumerate(test_overall_data, 3):
        for col, value in enumerate(row_data, 1):
            cell = ws_test_overall.cell(row=row, column=col, value=value)
            cell.border = border
            if row == 3:  # Header row
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
            elif col == 2 and isinstance(value, float):
                cell.number_format = '0.0000'
                cell.alignment = center_align
    
    ws_test_overall.column_dimensions['A'].width = 25
    ws_test_overall.column_dimensions['B'].width = 20
    
    ws_test_per_class = wb.create_sheet("Test Set - Per-Class", 2)
    ws_test_per_class['A1'] = f"Per-Class Test Metrics - {model_type.upper()}"
    ws_test_per_class['A1'].font = Font(bold=True, size=12)
    
    per_class_headers = ['Class', 'Gesture Name', 'Precision', 'Recall', 'F1-Score']
    for col, header in enumerate(per_class_headers, 1):
        cell = ws_test_per_class.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
    
    gesture_names = [
        "none", "open palm", "fist", "thumbs up", "peace", "OK", 
        "wave", "help signal", "swipe left", "swipe right"
    ]
    
    for class_idx in range(10):
        row = class_idx + 4
        row_data = [
            class_idx,
            gesture_names[class_idx],
            test_metrics['per_class_precision'][class_idx],
            test_metrics['per_class_recall'][class_idx],
            test_metrics['per_class_f1'][class_idx],
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws_test_per_class.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = center_align
            if isinstance(value, float):
                cell.number_format = '0.0000'
    
    ws_test_per_class.column_dimensions['A'].width = 10
    ws_test_per_class.column_dimensions['B'].width = 15
    ws_test_per_class.column_dimensions['C'].width = 12
    ws_test_per_class.column_dimensions['D'].width = 12
    ws_test_per_class.column_dimensions['E'].width = 12
    
    ws_confusion = wb.create_sheet("Confusion Matrix", 3)
    ws_confusion['A1'] = f"Confusion Matrix - {model_type.upper()}"
    ws_confusion['A1'].font = Font(bold=True, size=12)
    
    for col, gesture in enumerate(gesture_names, 2):
        cell = ws_confusion.cell(row=3, column=col, value=f"Pred: {gesture[:8]}")
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
    
    for row_idx, gesture in enumerate(gesture_names, 4):
        cell = ws_confusion.cell(row=row_idx, column=1, value=f"True: {gesture[:8]}")
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = center_align
        
        for col_idx in range(10):
            cell = ws_confusion.cell(row=row_idx, column=col_idx+2, value=int(confusion_mat[row_idx-4, col_idx]))
            cell.border = border
            cell.alignment = center_align
            if row_idx - 4 == col_idx:
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    ws_confusion.column_dimensions['A'].width = 15
    for col in range(2, 12):
        ws_confusion.column_dimensions[get_column_letter(col)].width = 12
    
    excel_path = metrics_path.replace('.metrics.csv', '.xlsx')
    wb.save(excel_path)
    print(f"\n✓ Excel report saved to: {excel_path}")
    return excel_path

# =========================
# train
# =========================


def cmd_train(args):
    args.out = f"model_{args.model_type}_seq{args.seq_len}.pth"
    
    rows = load_index(args.index)
    if not rows:
        raise RuntimeError("No data rows found in index.csv")

    # First split: 80% train+val, 20% test
    X_temp, X_test = train_test_split(
        rows,
        test_size=0.2,
        random_state=42,
        stratify=[r[1] for r in rows]
    )
    
    X_train, X_val = train_test_split(
        X_temp,
        test_size=0.25,
        random_state=42,
        stratify=[r[1] for r in X_temp]
    )

    print(f"\nData split: Train={len(X_train)} ({len(X_train)/len(rows)*100:.1f}%), "
          f"Val={len(X_val)} ({len(X_val)/len(rows)*100:.1f}%), "
          f"Test={len(X_test)} ({len(X_test)/len(rows)*100:.1f}%)\n")

    tr_ds = SeqDS(X_train, seq_len=args.seq_len)
    va_ds = SeqDS(X_val, seq_len=args.seq_len)
    te_ds = SeqDS(X_test, seq_len=args.seq_len)
    tr_dl = DataLoader(tr_ds, batch_size=args.batch, shuffle=True, num_workers=0)
    va_dl = DataLoader(va_ds, batch_size=args.batch, shuffle=False, num_workers=0)
    te_dl = DataLoader(te_ds, batch_size=args.batch, shuffle=False, num_workers=0)

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    model = make_model(args.model_type, seq_len=args.seq_len).to(device)
    opt = optim.AdamW(model.parameters(), lr=args.lr)
    crit = nn.CrossEntropyLoss()

    out_dir = os.path.dirname(args.out)
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir, exist_ok=True)

    best_f1 = -1.0
    best_epoch = None
    best_model_path = args.out + ".best.pth"
    
    train_metrics_history = []
    val_metrics_history = []

    metrics_path = args.out + ".metrics.csv"
    with open(metrics_path, "w") as mf:
        mf.write("epoch,model_type,train_acc,val_accuracy,val_macro_f1,val_precision,val_recall,saved\n")

        for ep in range(1, args.epochs + 1):
            model.train()
            tot = 0
            correct = 0
            for X, y in tr_dl:
                X, y = X.to(device), y.to(device)
                logits = model(X)
                loss = crit(logits, y)
                opt.zero_grad()
                loss.backward()
                opt.step()
                correct += (logits.argmax(1) == y).sum().item()
                tot += y.size(0)
            tr_acc = correct / tot if tot > 0 else 0.0

            # ---------- validate ----------
            model.eval()
            all_preds = []
            all_labels = []
            tp = [0] * NUM_CLASSES
            fp = [0] * NUM_CLASSES
            fn = [0] * NUM_CLASSES
            
            with torch.no_grad():
                for X, y in va_dl:
                    X, y = X.to(device), y.to(device)
                    pred = model(X).argmax(1)
                    all_preds.extend(pred.cpu().numpy())
                    all_labels.extend(y.cpu().numpy())
                    
                    for c in range(NUM_CLASSES):
                        tp[c] += ((pred == c) & (y == c)).sum().item()
                        fp[c] += ((pred == c) & (y != c)).sum().item()
                        fn[c] += ((pred != c) & (y == c)).sum().item()

            val_accuracy = accuracy_score(all_labels, all_preds)

            precisions = []
            recalls = []
            f1s = []
            
            for c in range(NUM_CLASSES):
                precision = tp[c] / (tp[c] + fp[c] + 1e-9)
                recall = tp[c] / (tp[c] + fn[c] + 1e-9)
                f1 = 2 * precision * recall / (precision + recall + 1e-9)
                precisions.append(precision)
                recalls.append(recall)
                f1s.append(f1)
            
            macro_f1 = float(np.mean(f1s))
            macro_precision = float(np.mean(precisions))
            macro_recall = float(np.mean(recalls))

            saved_flag = 0
            if macro_f1 > best_f1:
                best_f1 = macro_f1
                best_epoch = ep
                torch.save(model.state_dict(), best_model_path)
                saved_flag = 1
                print(f"  [checkpoint] saved best model (epoch {ep}, macroF1={macro_f1:.3f})")

            print(f"Epoch {ep:02d} | {args.model_type} | "
                  f"train_acc {tr_acc:.3f} | val_acc {val_accuracy:.3f} | "
                  f"val_macroF1 {macro_f1:.3f} | prec {macro_precision:.3f} | recall {macro_recall:.3f}")

            mf.write(f"{ep},{args.model_type},{tr_acc:.6f},{val_accuracy:.6f},{macro_f1:.6f},{macro_precision:.6f},{macro_recall:.6f},{saved_flag}\n")
            mf.flush()
            
            train_metrics_history.append({'accuracy': tr_acc})
            val_metrics_history.append({
                'accuracy': val_accuracy,
                'macro_f1': macro_f1,
                'macro_precision': macro_precision,
                'macro_recall': macro_recall
            })

    final_out = args.out
    torch.save(model.state_dict(), final_out)
    
    model.load_state_dict(torch.load(best_model_path, map_location=device))
    model.eval()
    
    test_all_preds = []
    test_all_labels = []
    test_tp = [0] * NUM_CLASSES
    test_fp = [0] * NUM_CLASSES
    test_fn = [0] * NUM_CLASSES
    
    with torch.no_grad():
        for X, y in te_dl:
            X, y = X.to(device), y.to(device)
            pred = model(X).argmax(1)
            test_all_preds.extend(pred.cpu().numpy())
            test_all_labels.extend(y.cpu().numpy())
            
            for c in range(NUM_CLASSES):
                test_tp[c] += ((pred == c) & (y == c)).sum().item()
                test_fp[c] += ((pred == c) & (y != c)).sum().item()
                test_fn[c] += ((pred != c) & (y == c)).sum().item()
    
    test_accuracy = accuracy_score(test_all_labels, test_all_preds)
    
    test_precisions = []
    test_recalls = []
    test_f1s = []
    
    for c in range(NUM_CLASSES):
        precision = test_tp[c] / (test_tp[c] + test_fp[c] + 1e-9)
        recall = test_tp[c] / (test_tp[c] + test_fn[c] + 1e-9)
        f1 = 2 * precision * recall / (precision + recall + 1e-9)
        test_precisions.append(precision)
        test_recalls.append(recall)
        test_f1s.append(f1)
    
    test_macro_f1 = float(np.mean(test_f1s))
    test_macro_precision = float(np.mean(test_precisions))
    test_macro_recall = float(np.mean(test_recalls))
    
    cm_test = confusion_matrix(test_all_labels, test_all_preds)
    
    test_metrics = {
        'accuracy': test_accuracy,
        'macro_f1': test_macro_f1,
        'macro_precision': test_macro_precision,
        'macro_recall': test_macro_recall,
        'per_class_precision': test_precisions,
        'per_class_recall': test_recalls,
        'per_class_f1': test_f1s,
    }
    
    save_metrics_to_excel(metrics_path, args.model_type, best_epoch, train_metrics_history, 
                          val_metrics_history, test_metrics, cm_test)
    
    print(f"\n" + "="*70)
    print("TRAINING COMPLETE")
    print("="*70)
    print(f"Best model saved to: {best_model_path} (epoch {best_epoch})")
    print(f"Final epoch model saved to: {final_out}")
    print(f"Metrics log saved to: {metrics_path}")
    print(f"Excel report saved to: {metrics_path.replace('.metrics.csv', '.xlsx')}")
    print(f"\nUse this model for inference: python gesture_temporal.py infer --model {best_model_path} --model_type {args.model_type} --seq_len {args.seq_len}")
    print("="*70)
# =========================
# infer
# =========================
def cmd_infer(args):
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    model = make_model(args.model_type, seq_len=args.seq_len).to(device)
    model.load_state_dict(torch.load(args.model, map_location=device))
    model.eval()

    cap = cv2.VideoCapture(0)
    cap.set(cv2.CAP_PROP_FRAME_WIDTH, args.width)
    cap.set(cv2.CAP_PROP_FRAME_HEIGHT, args.height)

    hands = mp_hands.Hands(
        static_image_mode=False,
        max_num_hands=2,
        model_complexity=1,
        min_detection_confidence=0.5,
        min_tracking_confidence=0.5,
    )

    window = deque(maxlen=args.seq_len)
    over_cnt = 0
    last_class = None
    last_fire = 0

    CLASSES = [str(i) for i in range(10)]

    while True:
        ok, frame = cap.read()
        if not ok:
            break
        if args.flip:
            frame = cv2.flip(frame, 1)

        rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        res = hands.process(rgb)

        pairs = []
        if res.multi_hand_landmarks:
            for lm, hd in zip(res.multi_hand_landmarks, res.multi_handedness):
                pairs.append((lm, hd))

        feat = handpack(pairs)
        window.append(feat)
        H, W = frame.shape[:2]

        if len(window) == args.seq_len:
            X = torch.from_numpy(np.stack(list(window))[None, :, :]).float().to(device)
            with torch.no_grad():
                prob = torch.softmax(model(X), dim=1)[0].cpu().numpy()

            # show per-class probabilities
            y0 = 20
            for i, c in enumerate(CLASSES):
                name = GESTURE_NAMES[i]
                text = f"{c}: {name[:8]} {prob[i]:.2f}"
                col = (0, 255, 0) if c != "0" else (200, 200, 200)
                cv2.putText(frame, text, (10, y0),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.5, col, 1)
                y0 += 20

            c_idx = int(prob.argmax())
            c_name = CLASSES[c_idx]

            if c_name != "0" and prob[c_idx] >= args.thresh:
                if last_class == c_idx:
                    over_cnt += 1
                else:
                    over_cnt = 1
                last_class = c_idx
            else:
                over_cnt = 0
                last_class = None

            if over_cnt >= args.hold:
                last_fire = time.time()
                over_cnt = 0
                gesture_label = GESTURE_NAMES[c_idx]
                last_label_text = f"{c_idx}: {gesture_label}"
                print(f"[DETECTED] {last_label_text}  (prob={prob[c_idx]:.3f})")

        # draw detection banner for ~1 second after trigger
        if time.time() - last_fire < 1.0 and last_label_text:
            overlay = frame.copy()
            cv2.rectangle(overlay, (0, 0), (W, 60), (0, 0, 255), -1)
            frame = cv2.addWeighted(overlay, 0.6, frame, 0.4, 0)
            cv2.putText(frame, f"DETECTED: {last_label_text}", (20, 40),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.9, (255, 255, 255), 2)

        cv2.imshow("infer_temporal", frame)
        if cv2.waitKey(1) & 0xFF in [27, ord('q')]:
            break

    hands.close()
    cap.release()
    cv2.destroyAllWindows()


# =========================
# main
# =========================
def main():
    ap = argparse.ArgumentParser(description="Temporal gesture pipeline (0..9) with TCN/LSTM/GRU/Transformer")
    sub = ap.add_subparsers(dest="cmd", required=True)

    # collect
    c = sub.add_parser("collect", help="Collect sequences with webcam")
    c.add_argument("--outdir", default="data")
    c.add_argument("--seq_len", type=int, default=SEQ_LEN_DEFAULT)
    c.add_argument("--width", type=int, default=1280)
    c.add_argument("--height", type=int, default=720)
    c.add_argument("--flip", action="store_true")
    c.set_defaults(func=cmd_collect)

    # train
    t = sub.add_parser("train", help="Train model")
    t.add_argument("--index", default="data/index.csv")
    t.add_argument("--seq_len", type=int, default=SEQ_LEN_DEFAULT)
    t.add_argument("--epochs", type=int, default=15)
    t.add_argument("--batch", type=int, default=64)
    t.add_argument("--lr", type=float, default=1e-3)
    t.add_argument("--out", default=None)
    t.add_argument("--model_type",
                   default="tcn",
                   choices=["tcn", "lstm", "gru", "transformer"])
    t.set_defaults(func=cmd_train)

    # infer
    i = sub.add_parser("infer", help="Live inference")
    i.add_argument("--model", default="model_tcn.pth")
    i.add_argument("--seq_len", type=int, default=SEQ_LEN_DEFAULT)
    i.add_argument("--width", type=int, default=1280)
    i.add_argument("--height", type=int, default=720)
    i.add_argument("--flip", action="store_true")
    i.add_argument("--thresh", type=float, default=0.7)
    i.add_argument("--hold", type=int, default=6)
    i.add_argument("--model_type",
                   default="tcn",
                   choices=["tcn", "lstm", "gru", "transformer"])
    i.set_defaults(func=cmd_infer)

    args = ap.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
